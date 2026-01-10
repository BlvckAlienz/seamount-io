"""
Excel Export Service for Bookkeeping Reports
"""
import io
import logging
from typing import Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class BookkeepingExporter:
    def export_trial_balance_to_excel(self, trial_balance: Dict, company_name: str = "Your Company") -> bytes:
        """Export trial balance to Excel format"""
        try:
            # Create a workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            
            # Company Header
            ws.merge_cells('A1:F1')
            ws['A1'] = company_name
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:F2')
            ws['A2'] = "Trial Balance Report"
            ws['A2'].font = Font(size=14)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Period
            ws.merge_cells('A3:F3')
            ws['A3'] = f"Period: {trial_balance.get('period_start', '')} to {trial_balance.get('period_end', '')}"
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Add some spacing
            ws.row_dimensions[4].height = 20
            
            # Column headers
            headers = ['Account Code', 'Account Name', 'Account Type', 'Debits', 'Credits', 'Balance']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(bottom=Side(style='thin'))
            
            # Add accounts data
            row = 6
            accounts = trial_balance.get('accounts', [])
            
            for account in accounts:
                ws.cell(row=row, column=1, value=account.get('account_code', ''))
                ws.cell(row=row, column=2, value=account.get('account_name', ''))
                ws.cell(row=row, column=3, value=account.get('account_type', ''))
                
                # Format amounts
                ws.cell(row=row, column=4, value=account.get('debits', 0))
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                
                ws.cell(row=row, column=5, value=account.get('credits', 0))
                ws.cell(row=row, column=5).number_format = '#,##0.00'
                
                ws.cell(row=row, column=6, value=account.get('balance', 0))
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                
                # Color negative balances
                if account.get('balance', 0) < 0:
                    ws.cell(row=row, column=6).font = Font(color="FF0000")
                
                row += 1
            
            # Add totals row
            ws.cell(row=row, column=3, value="TOTALS:").font = Font(bold=True)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            
            ws.cell(row=row, column=4, value=trial_balance.get('total_debits', 0))
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=4).font = Font(bold=True)
            
            ws.cell(row=row, column=5, value=trial_balance.get('total_credits', 0))
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=5).font = Font(bold=True)
            
            # Balance status
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            
            status = "✅ BALANCED" if trial_balance.get('is_balanced') else "⚠️ NOT BALANCED"
            color = "00B050" if trial_balance.get('is_balanced') else "FF0000"
            
            ws.cell(row=row, column=1, value=f"STATUS: {status}")
            ws.cell(row=row, column=1).font = Font(bold=True, color=color)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"✅ Excel report generated with {len(accounts)} accounts")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"❌ Excel export failed: {str(e)}")
            raise